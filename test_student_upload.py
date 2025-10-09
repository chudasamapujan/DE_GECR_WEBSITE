"""
Test Student Upload and Portal Functionality
Creates sample Excel file and tests student management flow
"""

import openpyxl
from datetime import datetime
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def create_sample_students_excel(filename='sample_students.xlsx'):
    """
    Create a sample students Excel file for testing bulk upload
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    
    # Create headers
    headers = ['Roll No', 'Name', 'Email', 'Password', 'Department', 'Semester', 'Phone']
    
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
    
    # Sample student data
    students = [
        ('2024001', 'Rahul Sharma', 'rahul.sharma@student.gecr.edu', 'rahul123', 'Computer Engineering', 5, '9876543210'),
        ('2024002', 'Priya Patel', 'priya.patel@student.gecr.edu', 'priya123', 'Computer Engineering', 5, '9876543211'),
        ('2024003', 'Amit Kumar', 'amit.kumar@student.gecr.edu', 'amit123', 'Computer Engineering', 5, '9876543212'),
        ('2024004', 'Sneha Desai', 'sneha.desai@student.gecr.edu', 'sneha123', 'AI & Data Science', 5, '9876543213'),
        ('2024005', 'Rohan Shah', 'rohan.shah@student.gecr.edu', 'rohan123', 'AI & Data Science', 5, '9876543214'),
    ]
    
    # Write student data
    for row_idx, student_data in enumerate(students, start=2):
        for col_idx, value in enumerate(student_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save workbook
    workbook.save(filename)
    print(f"✅ Sample Excel file created: {filename}")
    print(f"   Students: {len(students)}")
    print(f"   Columns: {', '.join(headers)}")
    
    return filename


def test_student_parser():
    """
    Test the student parser utility
    """
    from utils.student_parser import parse_students_excel
    
    # Create sample file
    filename = create_sample_students_excel()
    
    try:
        # Parse the file
        print("\n📊 Parsing Excel file...")
        result = parse_students_excel(filename)
        
        print(f"✅ Parsing completed!")
        print(f"   Students extracted: {result['total_students']}")
        print(f"   Errors: {len(result['errors'])}")
        
        if result['errors']:
            print("   ⚠️ Errors encountered:")
            for error in result['errors']:
                print(f"      - {error}")
        
        # Show sample students
        print(f"\n📝 Sample students (first 3):")
        for i, student in enumerate(result['students'][:3], 1):
            print(f"   {i}. {student['name']} ({student['roll_no']})")
            print(f"      Email: {student['email']}, Dept: {student['department']}, Sem: {student['semester']}")
        
        return result
        
    finally:
        # Cleanup
        if os.path.exists(filename):
            os.remove(filename)
            print(f"\n🧹 Cleaned up: {filename}")


def test_student_api():
    """
    Test student management API endpoints
    Requires Flask app to be running
    """
    import requests
    
    print("\n" + "=" * 70)
    print("TESTING STUDENT MANAGEMENT API")
    print("=" * 70)
    print("Note: Make sure Flask app is running on http://localhost:5000")
    
    session = requests.Session()
    base_url = 'http://localhost:5000'
    
    # Login as faculty
    print("\n🔐 Logging in as faculty...")
    login_response = session.post(f'{base_url}/api/auth/faculty/login',
        json={'email': 'faculty@gecr.edu', 'password': 'password'})
    
    if login_response.status_code != 200:
        print(f"   ❌ Login failed: {login_response.json()}")
        return
    
    print("   ✅ Login successful!")
    
    # Test 1: Add single student
    print("\n📝 Test 1: Adding single student...")
    add_response = session.post(f'{base_url}/api/faculty/students',
        json={
            'roll_no': '2025001',
            'name': 'Test Student API',
            'email': 'test.api@student.gecr.edu',
            'password': 'test123',
            'department': 'Computer Engineering',
            'semester': 5,
            'phone': '1234567890'
        })
    
    if add_response.status_code == 201:
        result = add_response.json()
        print(f"   ✅ Student added: {result['student']['name']}")
        print(f"      Roll No: {result['student']['roll_no']}")
    else:
        print(f"   ❌ Failed: {add_response.json()}")
    
    # Test 2: Upload Excel with students
    print("\n📤 Test 2: Uploading students from Excel...")
    filename = create_sample_students_excel('test_upload.xlsx')
    
    try:
        with open(filename, 'rb') as f:
            files = {'file': (filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            upload_response = session.post(f'{base_url}/api/faculty/students/upload', files=files)
        
        if upload_response.status_code == 200:
            result = upload_response.json()
            print(f"   ✅ Upload completed!")
            print(f"      Created: {result['created']}")
            print(f"      Skipped: {result['skipped']}")
            print(f"      Errors: {result['errors']}")
        else:
            print(f"   ❌ Failed: {upload_response.json()}")
    finally:
        if os.path.exists(filename):
            os.remove(filename)
    
    # Test 3: Get all students
    print("\n📋 Test 3: Fetching all students...")
    get_response = session.get(f'{base_url}/api/faculty/students')
    
    if get_response.status_code == 200:
        result = get_response.json()
        print(f"   ✅ Found {result['total']} students")
        for student in result['students'][:3]:
            print(f"      - {student['name']} ({student['roll_no']}) - Attendance: {student.get('attendance_percentage', 0)}%")
    else:
        print(f"   ❌ Failed: {get_response.json()}")
    
    # Test 4: Student login and profile
    print("\n👨‍🎓 Test 4: Student login and profile...")
    student_session = requests.Session()
    
    login_response = student_session.post(f'{base_url}/api/auth/student/login',
        json={'email': '2025001', 'password': 'test123'})  # Try roll number as email
    
    if login_response.status_code != 200:
        # Try with actual email
        login_response = student_session.post(f'{base_url}/api/auth/student/login',
            json={'email': 'test.api@student.gecr.edu', 'password': 'test123'})
    
    if login_response.status_code == 200:
        print("   ✅ Student logged in!")
        
        # Get profile
        profile_response = student_session.get(f'{base_url}/api/student/profile')
        if profile_response.status_code == 200:
            profile = profile_response.json()
            print(f"      Name: {profile.get('name')}")
            print(f"      Roll No: {profile.get('roll_no')}")
            print(f"      Enrolled Subjects: {len(profile.get('enrolled_subjects', []))}")
        
        # Get subjects
        subjects_response = student_session.get(f'{base_url}/api/student/subjects')
        if subjects_response.status_code == 200:
            subjects = subjects_response.json()
            print(f"      Total Subjects: {subjects.get('total_subjects', 0)}")
        
        # Get attendance
        attendance_response = student_session.get(f'{base_url}/api/student/attendance')
        if attendance_response.status_code == 200:
            attendance = attendance_response.json()
            print(f"      Overall Attendance: {attendance.get('overall_attendance_percentage', 0)}%")
    else:
        print(f"   ❌ Student login failed: {login_response.json()}")


if __name__ == '__main__':
    print("=" * 70)
    print("STUDENT MANAGEMENT - TEST SCRIPT")
    print("=" * 70)
    
    # Test 1: Parser only (no server needed)
    print("\n[TEST 1] Testing Student Excel Parser")
    print("-" * 70)
    test_student_parser()
    
    # Test 2: API endpoints (requires running server)
    print("\n" + "=" * 70)
    response = input("\nDo you want to test the API endpoints? (requires running Flask app) [y/N]: ")
    if response.lower() == 'y':
        test_student_api()
    else:
        print("Skipping API tests.")
    
    print("\n" + "=" * 70)
    print("✅ Testing completed!")
    print("=" * 70)
