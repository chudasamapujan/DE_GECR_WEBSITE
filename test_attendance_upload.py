"""
Test script for Excel attendance upload functionality
Creates a sample Excel file and tests the upload endpoint
"""

import openpyxl
from datetime import datetime, timedelta
import os

def create_sample_attendance_excel(filename='sample_attendance.xlsx'):
    """
    Create a sample attendance Excel file for testing
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    
    # Create headers
    headers = ['Roll No']
    base_date = datetime(2024, 1, 15)
    
    # Add 5 date columns
    for i in range(5):
        date = base_date + timedelta(days=i)
        headers.append(date.strftime('%d/%m/%Y'))
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
    
    # Sample student data with roll numbers
    students = [
        ('2020001', ['P', 'P', 'P', 'A', 'P']),
        ('2020002', ['P', 'A', 'P', 'P', 'L']),
        ('2020003', ['A', 'P', 'P', 'P', 'P']),
        ('2020004', ['P', 'P', 'L', 'P', 'P']),
        ('2020005', ['P', 'P', 'P', 'P', 'A']),
    ]
    
    # Write student data
    for row_idx, (roll_no, attendance) in enumerate(students, start=2):
        sheet.cell(row=row_idx, column=1, value=roll_no)
        for col_idx, status in enumerate(attendance, start=2):
            sheet.cell(row=row_idx, column=col_idx, value=status)
    
    # Save workbook
    workbook.save(filename)
    print(f"✅ Sample Excel file created: {filename}")
    print(f"   Headers: {headers}")
    print(f"   Students: {len(students)}")
    print(f"   Total records: {len(students) * 5}")
    
    return filename

def test_excel_parser():
    """
    Test the Excel parser utility
    """
    from utils.excel_parser import parse_attendance_excel
    
    # Create sample file
    filename = create_sample_attendance_excel()
    
    try:
        # Parse the file
        print("\n📊 Parsing Excel file...")
        result = parse_attendance_excel(filename)
        
        print(f"✅ Parsing completed!")
        print(f"   Dates found: {len(result['dates'])}")
        print(f"   Date values: {[d.strftime('%Y-%m-%d') for d in result['dates']]}")
        print(f"   Records extracted: {result['total_records']}")
        print(f"   Errors: {len(result['errors'])}")
        
        if result['errors']:
            print("   ⚠️ Errors encountered:")
            for error in result['errors']:
                print(f"      - {error}")
        
        # Show sample records
        print(f"\n📝 Sample records (first 3):")
        for i, record in enumerate(result['records'][:3], 1):
            print(f"   {i}. Roll: {record['student_roll_no']}, Date: {record['date']}, Status: {record['status']}")
        
        return result
        
    finally:
        # Cleanup
        if os.path.exists(filename):
            os.remove(filename)
            print(f"\n🧹 Cleaned up: {filename}")

def test_upload_endpoint():
    """
    Test the upload endpoint with a sample file
    Note: This requires the Flask app to be running
    """
    import requests
    
    # Create sample file
    filename = create_sample_attendance_excel()
    
    try:
        print("\n🚀 Testing upload endpoint...")
        print("   Note: Make sure Flask app is running on http://localhost:5000")
        
        # Login first to get session
        login_url = 'http://localhost:5000/api/auth/faculty/login'
        login_data = {
            'email': 'faculty@gecr.edu',  # Update with actual faculty email
            'password': 'password123'      # Update with actual password
        }
        
        session = requests.Session()
        
        print(f"\n🔐 Logging in as faculty...")
        login_response = session.post(login_url, json=login_data)
        
        if login_response.status_code != 200:
            print(f"   ❌ Login failed: {login_response.json()}")
            return
        
        print(f"   ✅ Login successful!")
        
        # Upload attendance
        upload_url = 'http://localhost:5000/api/faculty/attendance/upload'
        
        with open(filename, 'rb') as f:
            files = {'file': (filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {'subject_id': '1'}  # Update with actual subject ID
            
            print(f"\n📤 Uploading attendance Excel...")
            upload_response = session.post(upload_url, files=files, data=data)
        
        if upload_response.status_code == 200:
            result = upload_response.json()
            print(f"   ✅ Upload successful!")
            print(f"   Records inserted: {result.get('records_inserted', 0)}")
            print(f"   Records updated: {result.get('records_updated', 0)}")
            print(f"   Records skipped: {result.get('records_skipped', 0)}")
            print(f"   Dates: {result.get('dates', [])}")
        else:
            print(f"   ❌ Upload failed: {upload_response.status_code}")
            print(f"   Response: {upload_response.json()}")
        
    finally:
        if os.path.exists(filename):
            os.remove(filename)
            print(f"\n🧹 Cleaned up: {filename}")

if __name__ == '__main__':
    print("=" * 60)
    print("Excel Attendance Upload Test Script")
    print("=" * 60)
    
    # Test 1: Parser only (no server needed)
    print("\n[TEST 1] Testing Excel Parser")
    print("-" * 60)
    test_excel_parser()
    
    # Test 2: Full upload (requires running server)
    print("\n" + "=" * 60)
    print("[TEST 2] Testing Upload Endpoint")
    print("-" * 60)
    
    response = input("\nDo you want to test the upload endpoint? (requires running Flask app) [y/N]: ")
    if response.lower() == 'y':
        test_upload_endpoint()
    else:
        print("Skipping upload endpoint test.")
    
    print("\n" + "=" * 60)
    print("✅ Testing completed!")
    print("=" * 60)
